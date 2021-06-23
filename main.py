from ortools.linear_solver import pywraplp
import openpyxl
import math

wb_drone = openpyxl.load_workbook("Data\Drone.xlsx")
wb_order = openpyxl.load_workbook("Data\Order.xlsx")
wb_depot = openpyxl.load_workbook("Data\Depot.xlsx")

sheet_drone = wb_drone.active
sheet_order = wb_order.active
sheet_depot = wb_depot.active

# Các hằng số
n_drone = sheet_drone.max_row - 1
n_order = sheet_order.max_row - 1
n_depot = sheet_depot.max_row - 1

T = 5
M = 3
a_short = 100
a_swap = 1
a_Dist = 0.1
a_Bat = 0.1
B_PickUp = 1000
B_Stage = 200
CRad = 20
L = 5
n = 8
MinTimea = 1

O = {}
max_speed_r = {}
CurX = {}
CurY = {}
initT = {}
prepT = {}
size_o = {}
type_food = {}
cap_r = {}
OX = {}
OY = {}
DX = {}
DY = {}
M_OX = {}
M_OY = {}
M_DX = {}
M_DY = {}
M_EX = {}
M_EY = {}
action = {1: 1, 2: 2, 3: 3, 4:4}          # move, load, unload, charge
depot = []
drone = []
LocX_e = {}
LocY_e = {}
BatteryCap_r = {}
weight_r = {}
BatThresh_r = {}
food = [1, 2]

# Nạp các hằng số từ file excel
for i in range(1, n_drone + 1):
    drone.append(i)

for i in range(1, 5):
    depot.append(i)

cout = 1
for i in range(2, 2 + n_order):
    initT[cout] = float((sheet_order.cell(row = i, column = 8)).value)
    prepT[cout] = float((sheet_order.cell(row = i, column = 9)).value)
    O[cout] = [initT[cout], prepT[cout]]
    OX[cout] = float((sheet_order.cell(row = i, column = 2)).value)
    OY[cout] = float((sheet_order.cell(row = i, column = 3)).value)
    DX[cout] = float((sheet_order.cell(row = i, column = 4)).value)
    DY[cout] = float((sheet_order.cell(row = i, column = 5)).value)
    size_o[cout] = float((sheet_order.cell(row = i, column = 6)).value)
    type_food[cout] = float((sheet_order.cell(row = i, column = 6)).value)
    cout += 1

cout = 1
for i in range(2, 2 + n_drone):
    max_speed_r[cout] = float((sheet_drone.cell(row = i, column = 4)).value)
    CurX[cout] = float((sheet_drone.cell(row = i, column = 2)).value)
    CurY[cout] = float((sheet_drone.cell(row = i, column = 3)).value)
    cap_r[cout] = weight_r[cout] = float((sheet_drone.cell(row = i, column = 5)).value)
    BatteryCap_r[cout] = 90 * (weight_r[cout] + cap_r[cout])
    BatThresh_r[cout] = (250 / max_speed_r[cout]) * (weight_r[cout] + cap_r[cout])
    cout += 1

cout = 1
for i in range(2, 2 + n_depot):
    LocX_e[cout] = float((sheet_depot.cell(row = i, column = 2)).value)
    LocY_e[cout] = float((sheet_depot.cell(row = i, column = 3)).value)
    cout += 1

# Các hằng số M
for i in range(1, n_order + 1):
    if OX[i] > (500 - OX[i]):
        M_OX[i] = OX[i]
    else:
        M_OX[i] = 500 - OX[i]
    if OY[i] > (500 - OY[i]): M_OY[i] = OY[i]
    else: M_OY[i] = 500 - OY[i]
    if DX[i] > (500 - DX[i]): M_DX[i] = DX[i]
    else: M_DX[i] = 500 - DX[i]
    if DY[i] > (500 - DY[i]): M_DY[i] = DY[i]
    else: M_DY[i] = 500 - DY[i]

for i in range(1, n_depot + 1):
    if LocX_e[i] > (500 - LocX_e[i]):
        M_EX[i] = LocX_e[i]
    else:
        M_EX[i] = 500 - LocX_e[i]
    if LocY_e[i] > (500 - LocY_e[i]):
        M_EY[i] = LocY_e[i]
    else:
        M_EY[i] = 500 - LocY_e[i]

# Các hàm để đưa biến vào bộ giải
def bool3_to_solver(solver, arr, a, b, c):
    for i in a:
        for j in b:
            for k in c:
                arr[i, j, k] = solver.BoolVar('')

def bool2_to_solver(solver, arr, a, b):
    for i in a:
        for j in b:
            arr[i, j] = solver.BoolVar('')

def bool1_to_solver(solver, arr, a):
    for i in a:
        arr[i] = solver.BoolVar('')

def float2_to_solver(solver, arr, a, b):
    for i in a:
        for j in b:
            arr[i, j] = solver.NumVar(-999999, 999999, '')

def pfloat2_to_solver(solver, arr, a, b):
    for i in a:
        for j in b:
            arr[i, j] = solver.NumVar(0, 999999, '')


def pint2_to_solver(solver, arr, a, b):
    for i in a:
        for j in b:
            arr[i, j] = solver.IntVar(0, 999999, '')

def pfloat1_to_solver(solver, arr, a):
    for i in a:
        arr[i] = solver.NumVar(0, 999999, '')

def pfloat3_to_solver(solver, arr, a, b, c):
    for i in a:
        for j in b:
            for k in c:
                arr[i, j, k] = solver.NumVar(0, 999999, '')

def checkRO(r, RO, O_k):
    for x in O_k:
        if [r, x] in RO:
            return 0
    return 1

def checkRE(r, RE):
    for x in range(1, 5):
        if [r, x] in RE:
            return 0
    return 1

# Tập các biến sẽ thay đổi sau mỗi chu kì
Priority = {}
RO = []
RE = []
#drone
z = {}  # tập các hành động của drone tại t Z(r,a,t) 1, 2, 3 move load unload
z_load = {}  # Zload(r,o,t)
z_unload = {}  # z_unload(r,o,t)
z_transit = {}  # z_Transit(r,o,t)
z_food = {}  # z_food (r, f, t)
z_depot = {}
v_bat = {}
v_short = {}
x = {}  # tọa dộ của drone x(r,t)
y = {}  # tọa dộ của drone y(r,t)
x0 = {}  # vận tốc của drone x'(r,t)
y0 = {}  # vận tốc của drone y'(r,t)
xDist = {}  # khoảng cách drone đi trong t
yDist = {}  # khoảng cách drone đi trong t

#order
z_NStart = {}
Ready = {}  # Ready(o, t)
Lateness = {}
x_O = {}  # khoảng cách drone đến điểm lấy hàng xO(r, o, t)
y_O = {}
x_D = {}
y_D = {}
x_E = {}
y_E = {}
z_stage = {}
x_stage = {}
y_stage = {}
v_O = {}         # khoảng cách giữa drone và đơn hàng
vE = {}          # khoảng cách giữa drone và điểm sạc
Loaded = {}      # Loaded[r, o]

Delivered = {}
count_deliver = 0
R_IDE = []
Total_delay = 0

save_ztransit = {}
save_zload = {}
save_zunload = {}
save_z = {}
save_x = {}
save_y = {}
save_x0 = {}
save_y0 = {}
save_vbat = {}
save_zdepot = {}
save_zfood = {}
Loaded_o = {}

#Update Priorities
def update_priority(O_k, Loaded):
    for o in O_k:
        Priority[o] = 1
    for r in drone:
        O_r = []
        for o in O_k:
            if (Loaded[r, o] == 1): O_r.append(o)
        for i in range(0, len(O_r)):
            for j in range (i + 1, len(O_r)):
                if (initT[O_r[i]] < initT[O_r[j]]):
                    temp = O_r[i]
                    O_r[i] = O_r[j]
                    O_r[j] = temp
        i = 1
        for o in O_r:
            Priority[o] = pow(cap_r[r], i)
            i += 1


def reset():
    global z, z_load, z_unload, z_transit, z_food, z_depot, z_NStart, v_bat, v_short, x, y, x0, y0, xDist, yDist, Lateness
    global x_O, y_O, x_D, y_D, x_E, y_E, z_stage, x_stage, y_stage, v_O, vE, Loaded, Ready
    z= {}  # tập các hành động của drone tại t Z(r,a,t) 1, 2, 3 move load unload
    z_load = {}     # Zload(r,o,t)
    z_unload = {}   # z_unload(r,o,t)
    z_transit = {}  # z_Transit(r,o,t)
    z_food = {}     # z_food (r, f, t)
    z_depot = {}
    z_NStart = {}
    v_bat = {}
    v_short = {}
    x = {}          # tọa dộ của drone x(r,t)
    y = {}          # tọa dộ của drone y(r,t)
    x0 = {}         # vận tốc của drone x'(r,t)
    y0 = {}         # vận tốc của drone y'(r,t)
    xDist = {}      # khoảng cách drone đi trong t
    yDist = {}      # khoảng cách drone đi trong t
    Ready = {}      # Ready(o, t)
    Lateness = {}
    x_O = {}         # khoảng cách drone đến điểm lấy hàng xO(r, o, t)
    y_O = {}
    x_D = {}
    y_D = {}
    x_E = {}
    y_E = {}
    z_stage = {}
    x_stage = {}
    y_stage = {}
    v_O = {}
    vE = {}

Load2 = {}
def solve(O_k, T_k, K, state):
    reset()
    print("\nloop %d" %(K))
    print(RO)
    print(O_k)
    # print(RE)
    solver = pywraplp.Solver.CreateSolver('SCIP')

    for i in RO:
        v_O[i[0], i[1]] = solver.NumVar(0, 999999, '')

    for i in RE:
        vE[i[0], i[1]] = solver.NumVar(0, 999999, '')

    bool3_to_solver(solver, z, drone, action, T_k)
    bool3_to_solver(solver, z_load, drone, O_k, T_k)
    bool3_to_solver(solver, z_unload, drone, O_k, T_k)
    bool3_to_solver(solver, z_transit, drone, O_k, T_k)
    bool3_to_solver(solver, z_food, drone, food, T_k)
    bool3_to_solver(solver, z_depot, drone, depot, T_k)
    bool1_to_solver(solver, z_stage, O_k)
    bool1_to_solver(solver, z_NStart, O_k)
    float2_to_solver(solver, x0, drone, T_k)
    float2_to_solver(solver, y0, drone, T_k)
    float2_to_solver(solver, x, drone, T_k)
    float2_to_solver(solver, y, drone, T_k)
    pfloat2_to_solver(solver, xDist, drone, T_k)
    pfloat2_to_solver(solver, yDist, drone, T_k)
    pint2_to_solver(solver, v_bat, drone, T_k)
    pint2_to_solver(solver, v_short, drone, T_k)
    pfloat1_to_solver(solver, x_stage, O_k)
    pfloat1_to_solver(solver, y_stage, O_k)
    pfloat3_to_solver(solver, x_O, drone, O_k, T_k)
    pfloat3_to_solver(solver, y_O, drone, O_k, T_k)
    pfloat3_to_solver(solver, x_D, drone, O_k, T_k)
    pfloat3_to_solver(solver, y_D, drone, O_k, T_k)
    pfloat3_to_solver(solver, x_E, drone, depot, T_k)
    pfloat3_to_solver(solver, y_E, drone, depot, T_k)

    # cập nhật trạng thái
    delarray = []
    for i in Loaded:
        delarray.append(i)
    for a in delarray:
        del Loaded[a]
    if state == 0:
        for r in drone:
            solver.Add(x[r, T_k[0]] == CurX[r])
            solver.Add(y[r, T_k[0]] == CurY[r])
            solver.Add(v_bat[r, T_k[0]] == BatteryCap_r[r])
            for o in O_k:
                solver.Add(z_transit[r, o, T_k[0]] == 0)
                Loaded[r,o] = 0
        # else:
        #     for r in drone:  # cập nhật tọa độ của drone ở chu kì trước
        #         solver.Add(x[r, T_k[0]] == save_x[r])
        #         solver.Add((y[r, T_k[0]] == save_y[r]))
        #         solver.Add(v_bat[r, T_k[0]] == save_vbat[r])
        #         for o in O_k:
        #             Loaded[r, o] = 0
        #             solver.Add(z_transit[r, o, T_k[0]] == 0)
    else:
        for r in drone:              # cập nhật trạng thái load của đơn hàng
            for o in O_k:
                if ((r, o, T_k[0]) in save_ztransit) and ((r, o, T_k[0]) in save_zload):
                    if ((save_ztransit[r, o, T_k[0]] >= 1) or (save_zload[r, o, T_k[0]] == 1) or (save_zunload[r, o, T_k[0]] == 1) ):
                        Loaded[r, o] = 1
                    else:
                        Loaded[r, o] = 0
                else:
                    Loaded[r, o] = 0

                    # if (save_ztransit[r, o, T_k[0]] == 1):
                #         Load2[r, o] = 1
                # if (r, o) in Load2:
                #     Loaded[r, o] = 1

        for r in drone:          # cập nhật tọa độ của drone ở chu kì trước
            solver.Add(x[r, T_k[0]] == save_x[r])
            solver.Add((y[r, T_k[0]] == save_y[r]))
            solver.Add(v_bat[r, T_k[0]] == save_vbat[r])
            # solver.Add(v_bat[r, T_k[0]] == BatteryCap_r[r])
            for a in range(1, 5):          # trạng thái của drone ở chu kì trước
                solver.Add(z[r, a, T_k[0]] == save_z[r, a])
            for i in range(1, 5):
                solver.Add(z_depot[r, i, T_k[0]] == save_zdepot[r, i, T_k[0]])
            for o in O_k:
                # if ((r, o, T_k[T - 1] - M) in save_ztransit) and ((r, o, T_k[T - 1] - M) in save_zload) and ((r, o, T_k[T - 1] - M) in save_zunload):
                #     for t in range(T_k[0], T_k[T - 1] - M + 1):
                #         # solver.Add(z_transit[r, o, t] == save_ztransit[r, o, t])
                #         solver.Add(z_load[r, o, t] == save_zload[r, o, t])
                #         solver.Add(z_unload[r, o, t] == save_zunload[r, o, t])
                if ((r, o, T_k[0]) in save_ztransit) and ((r, o, T_k[0]) in save_zload) and ((r, o, T_k[0]) in save_zunload):
                    solver.Add(z_transit[r, o, T_k[0]] == save_ztransit[r, o, T_k[0]])
                    solver.Add(z_load[r, o, T_k[0]] == save_zload[r, o, T_k[0]])
                    solver.Add(z_unload[r, o, T_k[0]] == save_zunload[r, o, T_k[0]])
                else:
                    solver.Add(z_transit[r, o, T_k[0]] == 0)
                    solver.Add(z_load[r, o, T_k[0]] == 0)
                    solver.Add(z_unload[r, o, T_k[0]] == 0)
            # for t in range(T_k[T - M - 2], T_k[T - 1] + 1):
            #     solver.Add(z[r, 1, t] == 1)
            #     for o in O_k:
            #         if ((r, o, T_k[T - M - 2]) in save_ztransit):
            #             solver.Add(z_transit[r, o, t] == (save_ztransit[r, o, T_k[T - M - 2]] + save_zload[r, o, T_k[T - M - 2]]))
            #         else:
            #             solver.Add(z_transit[r, o, t] == 0)
            #     for e in range(1, 5):
            #         if (r, e, t) in save_zdepot:
            #             solver.Add(z_depot[r, e, t] == save_zdepot[r, e, T_k[T - M - 2]])
            #     for f in range(1, 3):
            #         if (r, f, t) in save_zfood:
            #             solver.Add(z_food[r, f, t] == save_zfood[r, f, T_k[T - M - 2]])

    # ràng buộc 1 an toàn
    for j in O_k:
        for k in T_k:
            Lateness[j, k] = k - initT[j] - prepT[j]

    # ràng buộc 2 an toàn
    for i in O_k:
        for j in T_k:
            if j >= (initT[i] + prepT[i]):
                Ready[i, j] = 1
            else:
                Ready[i, j] = 0

    # ràng buộc 5 an toàn
    for i in drone:
        for j in T_k:
            solver.Add(solver.Sum([z[i, k, j] for k in range(1, 5)]) == 1)

    # ràng buộc 6 an toàn
    for i in drone:
        for j in O_k:
            for k in T_k:
                solver.Add(z_load[i, j, k] <= z[i, 2, k])

    # ràng buộc 7 an toàn
    for i in drone:
        for j in O_k:
            for k in T_k:
                solver.Add(z_unload[i, j, k] <= z[i, 3, k])

    # ràng buộc 8

    # ràng buộc 9 an toàn
    for i in drone:
        for j in T_k:
            solver.Add( (2/(2-math.sqrt(2)))*x0[i, j] + math.sqrt(2)*y0[i, j] <= 2 / (2 - math.sqrt(2)) * z[i,1,j] * max_speed_r[i])
            solver.Add( math.sqrt(2) * x0[i, j] + 2 / (2 - math.sqrt(2)) * y0[i, j] <= 2 / (2 - math.sqrt(2)) * z[i,1,j] * max_speed_r[i])
            solver.Add( -math.sqrt(2) * x0[i, j] + 2 / (2 - math.sqrt(2)) * y0[i, j] <= 2 / (2 - math.sqrt(2)) * z[i,1,j] * max_speed_r[i])
            solver.Add(-2 / (2 - math.sqrt(2)) * x0[i, j] + math.sqrt(2) * y0[i, j] <= 2 / (2 - math.sqrt(2)) * z[i,1,j] * max_speed_r[i])
            solver.Add(-2 / (2 - math.sqrt(2)) * x0[i, j] - math.sqrt(2) * y0[i, j] <= 2 / (2 - math.sqrt(2)) * z[i,1,j] * max_speed_r[i])
            solver.Add( -math.sqrt(2) * x0[i, j] - 2 / (2 - math.sqrt(2)) * y0[i, j]<= 2 / (2 - math.sqrt(2)) * z[i,1,j] * max_speed_r[i])
            solver.Add( math.sqrt(2) * x0[i, j] - 2 / (2 - math.sqrt(2)) * y0[i, j] <= 2 / (2 - math.sqrt(2)) * z[i,1,j] * max_speed_r[i])
            solver.Add(2 / (2 - math.sqrt(2)) * x0[i, j] - math.sqrt(2) * y0[i, j] <= 2 / (2 - math.sqrt(2)) * z[i,1,j] * max_speed_r[i])

    # ràng buộc 10 11 an toàn
    for i in drone:
        for j in T_k:
            if j == T_k[0]: continue
            solver.Add(x[i, j] == (x[i, j - 1] + x0[i, j]))
            solver.Add(y[i, j] == (y[i, j - 1] + y0[i, j]))

    # ràng buộc 12 13 14 15
    for i in drone:
        for j in T_k:
            solver.Add(xDist[i, j] >= x0[i, j])
            solver.Add(xDist[i, j] >= -x0[i, j])
            solver.Add(yDist[i, j] >= y0[i, j])
            solver.Add(yDist[i, j] >= -y0[i, j])

    # ràng buộc 16
    for j in O_k:
        for k in T_k:
            solver.Add(solver.Sum([z_transit[i, j, k] for i in drone]) <= 1)

    # ràng buộc 17 an toàn
    for i in drone:
        for j in O_k:
            for k in T_k:
                solver.Add(z_load[i, j, k] <= Ready[j, k])

    # ràng buộc 18 - 27
    for i in drone:
        for j in O_k:
            for k in T_k:
                solver.Add(x_O[i, j, k] >= (x[i, k] - OX[j]))
                solver.Add(x_O[i, j, k] >= (-x[i, k] + OX[j]))
                solver.Add(y_O[i, j, k] >= (y[i, k] - OY[j]))
                solver.Add(y_O[i, j, k] >= (-y[i, k] + OY[j]))
                solver.Add((x_O[i, j, k] + y_O[i, j, k] + (M_OX[j] + M_OY[j])*z_load[i,j,k]) <= (M_OX[j] + M_OY[j]))

                solver.Add(x_D[i, j, k] >= (x[i, k] - DX[j]))
                solver.Add(x_D[i, j, k] >= (-x[i, k] + DX[j]))
                solver.Add(y_D[i, j, k] >= (y[i, k] - DY[j]))
                solver.Add(y_D[i, j, k] >= (-y[i, k] + DY[j]))
                solver.Add((x_D[i, j, k] + y_D[i, j, k] + (M_DX[j] + M_DY[j]) * z_unload[i, j, k]) <= (M_DX[j] + M_DY[j]))

    # ràng buộc 28
    for i in drone:
        for k in T_k:
            solver.Add(solver.Sum([(z_transit[i, j, k]*size_o[j]) for j in O_k]) <= cap_r[i])

    # ràng buộc 29
    for j in O_k:
        if type_food[j] == 1:
            for i in drone:
                for k in T_k:
                    solver.Add(z_transit[i, j, k] <= z_food[i,1,k])

    for j in O_k:
        if type_food[j] == 2:
            for i in drone:
                for k in T_k:
                    solver.Add(z_transit[i, j , k] <= z_food[i, 2, k])

    # ràng buộc 30
    for i in drone:
        for k in T_k:
            solver.Add((z_food[i, 1, k] + z_food[i, 2, k]) <= 1)

    # ràng buộc 31
    for i in drone:
        for j in O_k:
            for k in T_k:
                if k == T_k[0]: continue
                solver.Add(z_transit[i, j, k] == (z_transit[i, j, k - 1] + z_load[i, j, k - 1] - z_unload[i, j, k]))

    # ràng buộc 32 33 34 35 36
    for i in drone:
        for e in range(1, 5):
            for k in T_k:
                solver.Add(x_E[i, e, k] >= x[i, k] - LocX_e[e])
                solver.Add(x_E[i, e, k] >= -x[i, k] + LocX_e[e])
                solver.Add(y_E[i, e, k] >= y[i, k] - LocY_e[e])
                solver.Add(y_E[i, e, k] >= -y[i, k] + LocY_e[e])
                solver.Add((x_E[i, e, k] + y_E[i, e, k] + (M_EX[e] + M_EY[e])*z_depot[i, e, k]) <= (M_EX[e] + M_EY[e] + CRad) )

    # ràng buộc 37
    for r in drone:
        for t in T_k:
            solver.Add(z[r, 4, t] <= solver.Sum([z_depot[r, e, t] for e in range(1, n_depot + 1)]))

    # ràng buộc 38 39
    for r in drone:
        for t in range(T_k[1], T_k[T-1] + 1):
            solver.Add(v_bat[r, t] <= (v_bat[r, t - 1] - weight_r[r]*z[r,1,t] + BatteryCap_r[r]*z[r,4,t] -
                                       solver.Sum([size_o[o]*z_transit[r, o, t] for o in O_k])))
            solver.Add(v_bat[r, t] <= BatteryCap_r[r])

    # ràng buộc 45 phải tính Loaded r,o trước nha
    for o in O_k:
        solver.Add((solver.Sum([z_load[r, o, t] for r in drone for t in range (T_k[1], T_k[T -1] + 1)]) + solver.Sum([Loaded[r, o] for r in drone])) <= 1)
        # solver.Add((solver.Sum([z_load[r, o, t] for r in drone for t in range (T_k[1], T_k[T - 1] + 1)]) + solver.Sum([Loaded[r, o] for r in drone])) <= 1)

    # ràng buộc 46
    t1 = T + (K-1)*M
    for o in O_k:
        solver.Add(solver.Sum([z_transit[r, o, T_k[T - 1]] for r in drone]) >= z_stage[o])

    # ràng buộc 47-50
    for o in O_k:
        for r in drone:
            solver.Add(x_stage[o] >= (x[r, T_k[T - 1]] - DX[o] - (1 - z_transit[r, o, T_k[T - 1]])*M_DX[o])/max_speed_r[r])
            solver.Add(x_stage[o] >= ((-x[r, T_k[T - 1]] + DX[o] - (1 - z_transit[r, o, T_k[T - 1]])*M_DX[o]))/max_speed_r[r])
            solver.Add(y_stage[o] >= ((y[r, T_k[T - 1]] - DY[o] - (1 - z_transit[r, o, T_k[T - 1]]) * M_DY[o])) / max_speed_r[r])
            solver.Add(y_stage[o] >= ((-y[r, T_k[T - 1]] + DY[o] - (1 - z_transit[r, o, T_k[T - 1]]) * M_DY[o])) / max_speed_r[r])

    # ràng buộc 51 52
    for o in O_k:
        solver.Add(z_NStart[o] + solver.Sum([Loaded[r, o] for r in drone ]) <= 1)
        # solver.Add(z_NStart[o] + solver.Sum([z_unload[r, o, t] for r in drone for t in T_k]) + solver.Sum([z_load[r, o, t] for r in drone for t in T_k]) + solver.Sum([z_transit[r, o, t] for r in drone for t in T_k]) >= 1)
        for r in drone:
            for t in T_k:
                solver.Add((z_NStart[o] + z_load[r, o, t]) <= 1)
                solver.Add((z_NStart[o] + z_transit[r, o, t]) <= 1)


    # ràng buộc 53
    for o in O_k:
        solver.Add((solver.Sum([z_unload[r, o, t] for r in drone
                                for t in T_k]) + z_stage[o] + z_NStart[o]) == 1)

    # ràng buộc 54
    t0 = 1 + (K - 1)*M + M
    for i in RO:
        solver.Add(v_O[i[0], i[1]] >= (x_O[i[0], i[1], t0] + y_O[i[0], i[1], t0] - (M_OX[i[1]] + M_OY[i[1]])*(1 - z_NStart[i[1]]) ))

    # ràng buộc 55
    for r in drone:
        for t in T_k:
            solver.Add(v_short[r, t] >= (BatThresh_r[r] - v_bat[r, t]))

    # ràng buộc 56
    t1 = 1 + (K - 1) * M + M
    for i in RE:
        solver.Add(vE[i[0], i[1]] >= (x_E[i[0], i[1], t1] + y_E[i[0], i[1], t1] - (M_EX[i[1]] +
                                    M_EY[i[1]]) *solver.Sum([z[i[0], 4, t] for t in T_k]) ))

    update_priority(O_k, Loaded)

    # hàm mục tiêu
    # solve
    objective = 0
    for r in drone:
        for t in T_k:
            objective += v_short[r, t] * a_short / cap_r[r]

    for i in RE:
        objective += a_swap * vE[i[0], i[1]] / max_speed_r[i[0]]

    for o in O_k:
        objective += (Lateness[o, T_k[T - 1]] + B_PickUp) * z_NStart[o]

    for o in O_k:
        objective += (Lateness[o, T_k[T - 1]] + B_Stage) * z_stage[o]

    for o in O_k:
        for r in drone:
            for t in T_k:
                objective += Lateness[o, t] * (z_load[r, o, t] + z_unload[r, o, t])

    for i in RO:
        objective += v_O[i[0], i[1]] / max_speed_r[i[0]]

    for o in O_k:
        objective += Priority[o]*(x_stage[o] + y_stage[o])

    for r in drone:
        for t in T_k:
            objective += a_Dist * (xDist[r, t] + yDist[r, t]) / max_speed_r[r]

    for r in drone:
        for t in T_k:
            objective += z[r, 4, t]

    for r in drone:
        for t in T_k:
            objective -= a_Bat*v_bat[r, t]/ cap_r[r]

    # print('Number of variables =', solver.NumVariables())
    # print('Number of constraints =', solver.NumConstraints())

    solver.Minimize(objective)

    status = solver.Solve()

    if status == pywraplp.Solver.OPTIMAL or status == pywraplp.Solver.FEASIBLE:
        print('Solution:')
        print('Objective value =', solver.Objective().Value())
        print("NStart")
        for a in z_NStart:
            print("%d : %d" %(a,z_NStart[a].solution_value()))

        print("stage")
        for a in z_stage:
            print("%d : %d" % (a, z_stage[a].solution_value()))

        # if K != 1:
            # print("stage %d" %(z_stage[2].solution_value()))
            # print(x_stage[2].solution_value())
            # for r in drone:
            #     for t in T_k:
            #         if z_unload[r, 2, t].solution_value() == 1:
            #             print("unload at %d" %(t))

        print ("load")
        for o in O_k:
            for r in drone:
                for t in T_k:
                    if (z_load[r,o,t].solution_value() == 1):
                        print("%d : %d" %(o, t))

        print("unload")
        for o in O_k:
            for r in drone:
                for t in T_k:
                    if (z_unload[r, o, t].solution_value() == 1):
                        print("%d : %d" % (o, t))

        global Total_delay
        for o in O_k:
            sum = 0
            for r in drone:
                for t in range(T_k[0], T_k[M]):
                    sum += z_unload[r, o, t].solution_value()
                    if z_unload[r, o, t].solution_value() == 1:
                        Total_delay += Lateness[o, t]
            if sum == 1:
                print("Deliver %d!" %(o))
                Delivered[o] = 1
                del O_k[O_k.index(o)]
                global count_deliver
                count_deliver += 1

        for r in drone:
            save_x[r] = x[r, T_k[M]].solution_value()
            save_y[r] = y[r, T_k[M]].solution_value()
            save_vbat[r] = v_bat[r, T_k[M]].solution_value()
            for a in range(1, 5):
                save_z[r, a] = z[r, a, T_k[M]].solution_value()
            for o in O_k:
                for t in T_k:
                    save_zunload[r, o, t] = z_unload[r, o, t].solution_value()
                    save_zload[r, o, t] = z_load[r, o, t].solution_value()
                    save_ztransit[r, o, t] = z_transit[r, o, t].solution_value()
            for t in T_k:
                for e in range(1, 5):
                    save_zdepot[r,e,t] = z_depot[r, e, t].solution_value()
                for i in range(1, 3):
                    save_zfood[r, i, t] = z_food[r, i, t].solution_value()

        global R_IDE
        R_IDE = []
        for r in drone:
            sum = 0
            for o in O_k:
                for t in T_k:
                    sum += z_load[r, o, t].solution_value() + z_transit[r, o, t].solution_value()
            if sum == 0:
                R_IDE.append(r)

#        update_RO(K, O_k, z_load, z_NStart, x, y)
        t_0 = 1 + (K - 1) * M
        t_1 = 1 + (K - 1) * M + M
        for o in O_k:
            if (z_NStart[o].solution_value() == 1):
                dmin = 999999
                r_target = -1
                for r in R_IDE:
                    if checkRO(r, RO, O_k):
                        d = abs(x[r, t_1].solution_value() - OX[o]) + abs(y[r, t_1].solution_value() - OY[o])
                        if (d <= dmin):
                            dmin = d
                            r_target = r
                if (r_target != -1):
                    RO.append([r_target, o])
        i = 0
        while (i < len(RO)):
            if RO[i][0] not in R_IDE:
                del RO[i]
                i -= 1
            i += 1
        i = 0
        while (i < len(RO)):
            sum = 0
            for r in drone:
                for t in range(t_0, t_1 + 1):
                    # sum += z_load[r, RO[i][1], t].solution_value()
                    sum += z_load[r, RO[i][1], t].solution_value() + z_unload[r, RO[i][1], t].solution_value()
            if sum >= 1:
                del RO[i]
                i -= 1
            i += 1

        for r in drone:
            sum = 0
            for t in range(t_1, t_0 + T):
                sum += v_short[r, t].solution_value()
            if ((checkRE(r, RE)) and (sum > 0)):
                dmin = 999999
                e_target = -1
                for e in range(1, 5):
                    d = abs(x[r, t_1].solution_value() - LocX_e[e]) + abs(y[r, t_1].solution_value() - LocY_e[e])
                    if (d <= dmin):
                        dmin = d
                        e_target = e
                if (e_target != -1):
                    RE.append([r, e_target])
        i = 0
        while (i < len(RE)):
            sum = 0
            for t in range(t_0, t_1 + 1):
                sum += z[RE[i][0], 4, t].solution_value()
            if sum >= 1:
                del RE[i]
                i -= 1
            i += 1

        return 1

    else:
        print('The problem does not have an optimal solution.')
        return 0

def dispatch():
    O_global = {}
    K = 1
    cout = 1
    status = 0
    while(count_deliver < n_order):
        O_k = []
        T_k = []
        while ((cout <= n_order) and (initT[cout] <= 1 + (K - 1)*M)):           # thêm vào các đơn hàng M phút trước
            O_global[cout] = cout
            Delivered[cout] = 0
            cout += 1
        for a in O_global:                 # Thiết lập O_k
            if ((initT[a] <= (1 + (K - 1)*M)) and ((initT[a] + prepT[a]) <= ((K - 1)*M + T + L)) and (Delivered[a] == 0)):
                O_k.append(a)
        for a in range(1, T + 1):          # thiết lập T_k
            T_k.append(a + (K - 1)*M)
        status = solve(O_k, T_k, K, status)
        K += 1
    print("Total delay = %d" %(Total_delay))

dispatch()

